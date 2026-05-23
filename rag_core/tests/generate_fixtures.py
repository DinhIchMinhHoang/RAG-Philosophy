"""
Script to generate test fixtures for TabularParser tests.
Run once to create test data files.

Usage:
    python -m rag_core.tests.generate_fixtures
"""

import os
import sys

# Add parent dir to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

import pandas as pd

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")
os.makedirs(FIXTURES_DIR, exist_ok=True)


def create_wide_table():
    """Generate Excel with 20 columns (wide table test)."""
    data = {
        f"Column_{i}": [f"Value_{j}_{i}" for j in range(100)]
        for i in range(20)
    }
    df = pd.DataFrame(data)
    path = os.path.join(FIXTURES_DIR, "wide_table_20cols.xlsx")
    df.to_excel(path, index=False)
    print(f"Created: {path}")


def create_normal_table():
    """Generate Excel with 10 columns (normal table test)."""
    data = {
        "Product": [f"Widget_{i}" for i in range(100)],
        "Price": [100 + i * 10 for i in range(100)],
        "Quantity": [50 - i for i in range(100)],
        "Supplier": [f"Supplier_{i % 5}" for i in range(100)],
        "Category": [f"Cat_{i % 3}" for i in range(100)],
        "InStock": [True if i % 2 == 0 else False for i in range(100)],
        "Rating": [round(3.5 + (i % 10) * 0.1, 1) for i in range(100)],
        "Reviews": [i * 7 for i in range(100)],
        "Shipping": [f"Day_{i % 7 + 1}" for i in range(100)],
        "Warehouse": [f"WH_{i % 3 + 1}" for i in range(100)],
    }
    df = pd.DataFrame(data)
    path = os.path.join(FIXTURES_DIR, "normal_table_10cols.xlsx")
    df.to_excel(path, index=False)
    print(f"Created: {path}")


def create_merged_cells():
    """Generate Excel with merged cells (merged cells test)."""
    num_rows = 100
    regions = ["North", "South", "East", "West"]
    quarters = ["Q1", "Q2"]
    
    data = {
        "Region": [regions[i % 4] for i in range(num_rows)],
        "Quarter": [quarters[i % 2] for i in range(num_rows)],
        "Month": [f"Month_{i % 12 + 1}" for i in range(num_rows)],
        "Revenue": [1000 + i * 50 for i in range(num_rows)],
        "Expenses": [500 + i * 20 for i in range(num_rows)],
        "Profit": [500 + i * 30 for i in range(num_rows)],
    }
    df = pd.DataFrame(data)
    
    path = os.path.join(FIXTURES_DIR, "merged_cells_basic.xlsx")
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
        ws = writer.sheets['Data']
        # Merge first 3 columns of header row (rows are 1-indexed in openpyxl)
        from openpyxl.utils import get_column_letter
        merge_range = f"A1:{get_column_letter(3)}1"
        ws.merge_cells(merge_range)
    print(f"Created: {path}")


def create_large_csv():
    """Generate CSV with 10000 rows (streaming test)."""
    data = {
        "ID": list(range(1, 10001)),
        "Name": [f"User_{i}" for i in range(1, 10001)],
        "Email": [f"user{i}@example.com" for i in range(1, 10001)],
        "Department": [f"Dept_{i % 20}" for i in range(1, 10001)],
        "Salary": [30000 + (i % 100) * 500 for i in range(1, 10001)],
    }
    df = pd.DataFrame(data)
    path = os.path.join(FIXTURES_DIR, "large_csv_10k.csv")
    df.to_csv(path, index=False)
    print(f"Created: {path}")


def create_special_chars_csv():
    """Generate CSV with special characters."""
    data = {
        "Text": [
            "Normal text",
            "Text with | pipe",
            "Text with \n newline",
            'Text with "quotes"',
            "Unicode: Ngày mai",
            "Tab\there",
            "Mixed: | and \n",
        ]
    }
    df = pd.DataFrame(data)
    path = os.path.join(FIXTURES_DIR, "special_chars.csv")
    df.to_csv(path, index=False)
    print(f"Created: {path}")


if __name__ == "__main__":
    print("Generating test fixtures...")
    try:
        create_wide_table()
        create_normal_table()
        create_merged_cells()
        create_large_csv()
        create_special_chars_csv()
        print("\nAll fixtures created successfully!")
    except ImportError as e:
        print(f"Missing dependency: {e}")
        print("Install with: pip install openpyxl pandas")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
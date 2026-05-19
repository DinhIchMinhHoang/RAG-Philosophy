"""
rag_core/parsers/tabular_parser.py - TabularParser V3 FINAL

Production-ready parser for Excel (.xlsx, .xls) and CSV files.
Handles:
- Merged cells via openpyxl (1:1 Grid Mapping)
- Wide table detection (>15 cols) → Key-Value fallback
- OOM protection via file size check and psutil MemoryGuard
- Explicit garbage collection after each sheet
- Citation contract compliance (page, source metadata)
- CPU thread limits (OMP_NUM_THREADS=1)
"""

from __future__ import annotations

import gc
import os

# Set thread limits BEFORE importing pandas/numpy
os.environ['OMP_NUM_THREADS'] = '1'
os.environ['MKL_NUM_THREADS'] = '1'
os.environ['OPENBLAS_NUM_THREADS'] = '1'

from typing import List, Optional

from langchain_core.documents import Document

try:
    from ..config import Config
    from ..common.logging_utils import get_logger
except ImportError:  # pragma: no cover
    from config import Config
    from common.logging_utils import get_logger

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import psutil
except ImportError:
    psutil = None

logger = get_logger(__name__)


class MemoryGuard:
    """Runtime memory protection using psutil."""
    
    @staticmethod
    def check_available_memory(min_required_mb: float = 100.0) -> bool:
        """Check if enough memory is available before loading file."""
        if psutil is None:
            return True  # Skip check if psutil not available
        
        available = psutil.virtual_memory().available / (1024 * 1024)
        if available < min_required_mb:
            logger.warning(
                f"MemoryGuard: insufficient memory. "
                f"Available: {available:.1f}MB, Required: {min_required_mb:.1f}MB"
            )
            return False
        return True


class TabularParser:
    """
    Unified parser for Excel and CSV files with production-grade safety features.
    
    Features:
    - Single I/O openpyxl loading for merged cells
    - Wide table detection (>15 cols) → Key-Value fallback
    - Row-based chunking (50 rows default)
    - Header prepending in every chunk
    - Citation contract: metadata.page, metadata.source
    - is_tabular flag for step2_chunker bypass
    """
    
    def __init__(
        self,
        max_rows: Optional[int] = None,
        max_cols: Optional[int] = None,
        max_size_mb: Optional[int] = None,
    ):
        self.max_rows = max_rows or getattr(Config, 'TABULAR_CHUNK_ROWS', 50)
        self.max_cols = max_cols or getattr(Config, 'TABULAR_MAX_COLS', 15)
        self.max_size_mb = max_size_mb or getattr(Config, 'MAX_TABULAR_SIZE_MB', 20)
        self._wb = None
    
    def _validate_file(self, file_path: str) -> None:
        """Validate file size and available memory before loading."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_size = os.path.getsize(file_path)
        if file_size > self.max_size_mb * 1024 * 1024:
            raise ValueError(
                f"File exceeds size limit: {file_size / 1024 / 1024:.1f}MB "
                f"(max: {self.max_size_mb}MB)"
            )
        
        # Estimate memory: file size * ~3 (openpyxl + pandas + overhead)
        estimated_mb = (file_size / (1024 * 1024)) * 3
        if not MemoryGuard.check_available_memory(min_required_mb=estimated_mb + 100):
            raise MemoryError(
                f"Insufficient memory to process file. "
                f"Required: ~{estimated_mb:.0f}MB"
            )
    
    def _df_to_documents(
        self,
        df,
        source: str,
        page_idx: int,
        sheet_name: str,
        doc_type: str,
        row_offset: int = 0,
    ) -> List[Document]:
        """
        Convert DataFrame chunk to LangChain Documents.
        
        Row-based chunking with header prepending.
        Wide table fallback to Key-Value format.
        """
        documents: List[Document] = []
        
        # Clean DataFrame: drop empty rows/cols, fill NaN
        df = df.dropna(how='all').dropna(axis=1, how='all').fillna("")
        if df.empty:
            return documents
        
        total_rows = len(df)
        is_wide_table = len(df.columns) > self.max_cols
        
        for chunk_start in range(0, total_rows, self.max_rows):
            chunk_end = min(chunk_start + self.max_rows, total_rows)
            chunk_df = df.iloc[chunk_start:chunk_end]
            
            # Format: Markdown Table or Key-Value (wide table fallback)
            if is_wide_table:
                lines = []
                for _, row in chunk_df.iterrows():
                    kv_pairs = [
                        f"{col}: {val}"
                        for col, val in row.items()
                        if str(val).strip()
                    ]
                    if kv_pairs:
                        lines.append(f"- {' | '.join(kv_pairs)}")
                data_content = "\n".join(lines)
            else:
                data_content = chunk_df.to_markdown(index=False)
            
            # Build document content with header context
            content = (
                f"Source: {sheet_name}\n"
                f"Rows: {row_offset + chunk_start + 1} to {row_offset + chunk_end}\n"
                f"Format: {'Key-Value (Wide Table)' if is_wide_table else 'Markdown Table'}\n\n"
                f"{data_content}"
            )
            
            documents.append(Document(
                page_content=content,
                metadata={
                    "source": source,
                    "page": page_idx,
                    "doc_type": doc_type,
                    "sheet_name": sheet_name,
                    "is_tabular": True,
                    "row_range": f"{row_offset + chunk_start + 1}-{row_offset + chunk_end}",
                }
            ))
        
        return documents
    
    def _parse_csv(self, file_path: str, source_name: str) -> List[Document]:
        """CSV: Streaming read with row offset tracking."""
        all_docs: List[Document] = []
        chunk_idx = 0
        row_offset = 0
        
        for chunk_df in pd.read_csv(file_path, chunksize=self.max_rows):
            docs = self._df_to_documents(
                df=chunk_df,
                source=source_name,
                page_idx=chunk_idx + 1,
                sheet_name=source_name,
                doc_type="tabular_csv",
                row_offset=row_offset,
            )
            all_docs.extend(docs)
            chunk_idx += 1
            row_offset += len(chunk_df)
        
        return all_docs
    
    def _parse_excel(self, file_path: str, source_name: str) -> List[Document]:
        """Excel: Single I/O with merged cell handling and explicit GC."""
        all_docs: List[Document] = []
        
        # Load workbook ONCE with data_only mode (single I/O, skip formulas)
        wb = None
        if openpyxl:
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                self._wb = wb
            except Exception as e:
                logger.warning(f"openpyxl unavailable for {source_name}: {e}")
        
        xls = None
        try:
            xls = pd.ExcelFile(file_path)
            
            for sheet_idx, sheet_name in enumerate(xls.sheet_names):
                # Read DataFrame (header=None for precise index mapping)
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                if df.empty:
                    continue
                
                # Patch merged cells: 1:1 Grid Mapping
                if wb and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for merged_range in ws.merged_cells.ranges:
                        min_col, min_row, max_col, max_row = merged_range.bounds
                        val = ws.cell(row=min_row, column=min_col).value
                        if val is not None:
                            # Map openpyxl coords (1-indexed) to pandas iloc (0-indexed)
                            df.iloc[min_row - 1:max_row, min_col - 1:max_col] = val
                
                # Auto-detect header row: find first row with ≥50% non-empty cells
                header_row = 0
                for i in range(min(10, len(df))):
                    non_empty = df.iloc[i].notna().sum()
                    if non_empty >= len(df.columns) * 0.5:
                        header_row = i
                        break

                # Promote header row to column names, clean NaN and whitespace
                df.columns = (
                    df.iloc[header_row]
                    .astype(str)
                    .str.strip()
                    .replace(r"^nan$", "", regex=True)
                )
                # Remove header row and all rows before it from data
                df = df.drop(index=range(header_row + 1)).reset_index(drop=True)
                
                # Convert to documents
                docs = self._df_to_documents(
                    df=df,
                    source=source_name,
                    page_idx=sheet_idx + 1,
                    sheet_name=sheet_name,
                    doc_type="tabular_excel",
                )
                all_docs.extend(docs)
                
                # Explicit garbage collection after each sheet
                del df
                gc.collect()
        
        finally:
            # Ensure workbook and ExcelFile cleanup
            if xls is not None:
                xls.close()
            if wb:
                wb.close()
                self._wb = None
        
        return all_docs
    
    def parse(self, file_path: str) -> List[Document]:
        """
        Main entry point - auto-detect format and parse.
        
        Args:
            file_path: Absolute or relative path to file.
            
        Returns:
            List[Document]: LangChain Documents with tabular metadata.
        """
        self._validate_file(file_path)
        
        ext = os.path.splitext(file_path)[1].lower()
        source_name = os.path.basename(file_path)
        
        try:
            if ext == '.csv':
                return self._parse_csv(file_path, source_name)
            elif ext in ['.xlsx', '.xls']:
                return self._parse_excel(file_path, source_name)
            else:
                raise ValueError(f"Unsupported format: {ext}")
        except Exception as exc:
            logger.error(f"TabularParser failed: {source_name}: {exc}")
            raise
    
    def __del__(self):
        """Destructor: safety net for workbook cleanup."""
        if hasattr(self, '_wb') and self._wb is not None:
            try:
                self._wb.close()
            except Exception:
                pass


# Lazy import pandas to allow module-level OMP settings
import pandas as pd  # noqa: E402